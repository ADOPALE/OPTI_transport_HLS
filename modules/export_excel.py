"""
export_excel.py — Export Excel multi-jours des résultats transport
==================================================================

Produit un classeur (BytesIO) à partir des résultats de
`moteur_postes.optimiser_postes_jour`, un par jour.

Onglets :
  0. Lecture            — mode d'emploi + alerte validité
  1. Synthèse flotte    — véhicules par type et par jour
  2. Synthèse chauffeurs— postes / relève par jour
  3. Tournées véhicules — chronologie détaillée (chargé/à vide, km, contenants)
  4. Planning chauffeurs— un poste par ligne, horaires et contenu
  5. Planning quais     — passages à quai par site
  6. Flux transportés   — flux servis et la mission qui les porte
  7. Flux non servis    — flux NON planifiés + CONTRAINTE BLOQUANTE explicite
  8. Contrôles          — vérification des contraintes (pause, amplitude, relève…)
  9. Indicateurs        — KPI par jour
"""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- styles ---
_TITRE = Font(bold=True, size=14, color="1F3864")
_ENTETE = Font(bold=True, color="FFFFFF")
_FILL_ENTETE = PatternFill("solid", fgColor="1F3864")
_FILL_ALERTE = PatternFill("solid", fgColor="C00000")
_FILL_OK = PatternFill("solid", fgColor="C6EFCE")
_FILL_KO = PatternFill("solid", fgColor="FFC7CE")
_FILL_WARN = PatternFill("solid", fgColor="FFEB9C")
_CENTRE = Alignment(horizontal="center", vertical="center")
_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
_BORD = Border(*[Side(style="thin", color="D9D9D9")] * 4)


def _hhmm(m):
    try:
        m = float(m)
        return f"{int(m // 60):02d}:{int(round(m % 60)):02d}"
    except Exception:
        return ""


def _entete(ws, ligne, colonnes, depart=1):
    for j, titre in enumerate(colonnes, depart):
        c = ws.cell(row=ligne, column=j, value=titre)
        c.font = _ENTETE; c.fill = _FILL_ENTETE; c.alignment = _CENTRE; c.border = _BORD
    ws.freeze_panes = ws.cell(row=ligne + 1, column=1)


def _largeurs(ws, largeurs):
    for j, w in enumerate(largeurs, 1):
        ws.column_dimensions[get_column_letter(j)].width = w


def _titre(ws, texte):
    c = ws.cell(row=1, column=1, value=texte)
    c.font = _TITRE
    ws.cell(row=2, column=1, value="")


# =====================================================================

def construire_excel(resultats_par_jour, params_logistique=None):
    """
    resultats_par_jour : dict {nom_jour: resultat} (sortie optimiser_postes_jour).
    Renvoie un io.BytesIO prêt à télécharger.
    """
    wb = Workbook()
    jours = list(resultats_par_jour.keys())
    total_non_servis = sum(len(r.get("non_servis", [])) for r in resultats_par_jour.values())
    total_anomalies = sum(r.get("audit", {}).get("nb_anomalies", 0)
                          for r in resultats_par_jour.values())

    _onglet_lecture(wb.active, jours, total_non_servis, total_anomalies)
    _onglet_synthese_flotte(wb.create_sheet("1. Synthèse flotte"), resultats_par_jour)
    _onglet_synthese_chauffeurs(wb.create_sheet("2. Synthèse chauffeurs"), resultats_par_jour)
    _onglet_tournees(wb.create_sheet("3. Tournées véhicules"), resultats_par_jour)
    _onglet_planning_chauffeurs(wb.create_sheet("4. Planning chauffeurs"), resultats_par_jour)
    _onglet_quais(wb.create_sheet("5. Planning quais"), resultats_par_jour)
    _onglet_flux_transportes(wb.create_sheet("6. Flux transportés"), resultats_par_jour)
    _onglet_flux_non_servis(wb.create_sheet("7. Flux NON servis"), resultats_par_jour)
    _onglet_controles(wb.create_sheet("8. Contrôles"), resultats_par_jour, params_logistique or {})
    _onglet_indicateurs(wb.create_sheet("9. Indicateurs"), resultats_par_jour)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# --------------------------------------------------------------- 0. Lecture
def _onglet_lecture(ws, jours, total_non_servis, total_anomalies=0):
    ws.title = "0. Lecture"
    _titre(ws, "Dimensionnement transport — CHU de Nantes")
    lignes = [
        "", "Ce classeur présente le dimensionnement de la flotte et des postes chauffeurs",
        "calculé pour chaque jour de la semaine.", "",
        f"Jours calculés : {', '.join(jours)}", "",
        "VALIDITÉ DE LA SOLUTION",
    ]
    for i, t in enumerate(lignes, 1):
        ws.cell(row=2 + i, column=1, value=t)
    r = 2 + len(lignes) + 1
    if total_non_servis == 0 and total_anomalies == 0:
        c = ws.cell(row=r, column=1,
                    value="✓ Solution VALIDE : tous les flux du périmètre sont planifiés.")
        c.fill = _FILL_OK; c.font = Font(bold=True)
    else:
        c = ws.cell(row=r, column=1,
                    value=f"✗ Solution NON VALIDE : {total_non_servis} flux non servi(s), "
                          f"{total_anomalies} anomalie(s) technique(s). "
                          f"Voir l'onglet « 7. Flux NON servis » : chaque ligne indique la "
                          f"CONTRAINTE BLOQUANTE et, si c'est un paramètre réglable, comment la lever.")
        c.fill = _FILL_ALERTE; c.font = Font(bold=True, color="FFFFFF")
    ws.cell(row=r + 2, column=1, value="Onglets : 1-Flotte · 2-Chauffeurs · 3-Tournées détaillées · "
            "4-Planning chauffeurs · 5-Quais · 6-Flux transportés · 7-Flux non servis · "
            "8-Contrôles contraintes · 9-Indicateurs.")
    _largeurs(ws, [120])


# --------------------------------------------------------------- 1. Flotte
def _onglet_synthese_flotte(ws, res):
    _titre(ws, "Synthèse flotte — véhicules nécessaires par type et par jour")
    jours = list(res.keys())
    types = sorted({t for r in res.values() for t in r["nb_vehicules"].keys()})
    _entete(ws, 3, ["Type de véhicule"] + jours + ["Maxi semaine"])
    r0 = 4
    for i, t in enumerate(types):
        ws.cell(row=r0 + i, column=1, value=t).border = _BORD
        vals = []
        for j, jour in enumerate(jours, 2):
            v = res[jour]["nb_vehicules"].get(t, 0)
            vals.append(v)
            ws.cell(row=r0 + i, column=j, value=v).alignment = _CENTRE
        c = ws.cell(row=r0 + i, column=2 + len(jours), value=max(vals) if vals else 0)
        c.font = Font(bold=True); c.alignment = _CENTRE
    rt = r0 + len(types)
    ws.cell(row=rt, column=1, value="TOTAL").font = Font(bold=True)
    for j, jour in enumerate(jours, 2):
        c = ws.cell(row=rt, column=j, value=res[jour]["metriques"]["nb_vehicules_total"])
        c.font = Font(bold=True); c.alignment = _CENTRE; c.fill = _FILL_WARN
    pic_total = max((res[j]["metriques"]["nb_vehicules_total"] for j in jours), default=0)
    c = ws.cell(row=rt, column=2 + len(jours), value=pic_total)
    c.font = Font(bold=True); c.fill = _FILL_WARN; c.alignment = _CENTRE
    ws.cell(row=rt + 2, column=1,
            value="Flotte à dimensionner = maximum hebdomadaire par type (colonne « Maxi semaine »).")
    _largeurs(ws, [26] + [11] * len(jours) + [13])


# --------------------------------------------------------------- 2. Chauffeurs
def _onglet_synthese_chauffeurs(ws, res):
    _titre(ws, "Synthèse chauffeurs — postes (vacations) et relève par jour")
    jours = list(res.keys())
    _entete(ws, 3, ["Jour", "Nb postes", "Nb véhicules", "Postes en relève (2/véh)",
                    "Occupation moyenne", "Amplitude moyenne"])
    for i, jour in enumerate(jours):
        r = res[jour]; postes = r["postes"]
        from collections import Counter
        c = Counter(p.id_vehicule for p in postes)
        releve = sum(1 for v in c.values() if v >= 2) * 2
        amp = sum(p.amplitude for p in postes) / len(postes) if postes else 0
        row = 4 + i
        ws.cell(row=row, column=1, value=jour).border = _BORD
        ws.cell(row=row, column=2, value=r["metriques"]["nb_postes"]).alignment = _CENTRE
        ws.cell(row=row, column=3, value=r["metriques"]["nb_vehicules_total"]).alignment = _CENTRE
        ws.cell(row=row, column=4, value=releve).alignment = _CENTRE
        ws.cell(row=row, column=5, value=f"{r['metriques']['occupation_moyenne']:.0f} %").alignment = _CENTRE
        ws.cell(row=row, column=6, value=_hhmm(amp)).alignment = _CENTRE
    _largeurs(ws, [14, 12, 14, 24, 18, 18])


# --------------------------------------------------------------- 3. Tournées
def _onglet_tournees(ws, res):
    _titre(ws, "Tournées véhicules — chronologie détaillée (étape par étape)")
    cols = ["Jour", "Véhicule", "Poste", "Type véh.", "Début", "Fin", "Durée (min)",
            "Étape", "Site départ", "Site arrivée", "Charge", "Distance (km)",
            "Contenants", "Flux concernés", "Fonction support"]
    _entete(ws, 3, cols)
    row = 4
    libelle_etape = {"PRISE": "Prise de poste", "APPROCHE_VIDE": "Approche à vide",
                     "MISSION": "Mission (chargé)", "MARGE": "Marge inter-job",
                     "RETOUR_VIDE": "Retour à vide", "NETTOYAGE": "Désinfection",
                     "PAUSE": "Pause (dépôt)", "ATTENTE": "Attente dispo",
                     "DISPONIBLE": "Disponible", "FIN": "Clôture poste"}
    for jour in res:
        for p in sorted(res[jour]["postes"], key=lambda p: (p.id_vehicule, p.h_debut)):
            for e in p.etapes:
                ws.cell(row=row, column=1, value=jour)
                ws.cell(row=row, column=2, value=p.id_vehicule)
                ws.cell(row=row, column=3, value=p.id)
                ws.cell(row=row, column=4, value=p.v_type)
                ws.cell(row=row, column=5, value=_hhmm(e.h_debut))
                ws.cell(row=row, column=6, value=_hhmm(e.h_fin))
                ws.cell(row=row, column=7, value=round(e.duree, 1))
                ce = ws.cell(row=row, column=8, value=libelle_etape.get(e.type, e.type))
                ws.cell(row=row, column=9, value=e.site_debut or "")
                ws.cell(row=row, column=10, value=e.site_fin or "")
                charge = "à vide" if e.a_vide else ("chargé" if e.type == "MISSION" else "")
                ws.cell(row=row, column=11, value=charge)
                ws.cell(row=row, column=12, value=round(e.distance, 1) if e.distance else "")
                if e.type == "MISSION" and e.mission:
                    m = e.mission
                    ws.cell(row=row, column=13,
                            value=f"{m.nb_contenants} × {m.libelle} ({m.fill:.0%})")
                    ws.cell(row=row, column=14,
                            value=", ".join(str(c[0]) for c in m.composantes))
                    ws.cell(row=row, column=15, value=m.fonction_support)
                    if e.a_vide is False:
                        ce.fill = _FILL_OK
                elif e.type in ("APPROCHE_VIDE", "RETOUR_VIDE"):
                    ce.fill = _FILL_WARN
                elif e.type == "PAUSE":
                    ce.fill = _FILL_KO
                row += 1
    _largeurs(ws, [10, 16, 16, 12, 8, 8, 11, 18, 18, 18, 9, 12, 26, 18, 16])


# --------------------------------------------------------------- 4. Planning chauffeurs
def _onglet_planning_chauffeurs(ws, res):
    _titre(ws, "Planning chauffeurs — un poste = une vacation")
    cols = ["Jour", "Véhicule", "Poste", "Type véh.", "Prise", "Fin", "Amplitude",
            "Nb missions", "Tps chargé", "Tps à vide", "Pause", "Occupation"]
    _entete(ws, 3, cols)
    row = 4
    for jour in res:
        for p in sorted(res[jour]["postes"], key=lambda p: (p.id_vehicule, p.h_debut)):
            pause = sum(e.duree for e in p.etapes if e.type == "PAUSE")
            ws.cell(row=row, column=1, value=jour)
            ws.cell(row=row, column=2, value=p.id_vehicule)
            ws.cell(row=row, column=3, value=p.id)
            ws.cell(row=row, column=4, value=p.v_type)
            ws.cell(row=row, column=5, value=_hhmm(p.h_debut))
            ws.cell(row=row, column=6, value=_hhmm(p.h_fin))
            ws.cell(row=row, column=7, value=_hhmm(p.amplitude))
            ws.cell(row=row, column=8, value=len(p.missions)).alignment = _CENTRE
            ws.cell(row=row, column=9, value=f"{p.temps_charge():.0f}")
            ws.cell(row=row, column=10, value=f"{p.temps_vide():.0f}")
            ws.cell(row=row, column=11, value=f"{pause:.0f}")
            ws.cell(row=row, column=12, value=f"{p.occupation():.0%}").alignment = _CENTRE
            row += 1
    _largeurs(ws, [10, 16, 16, 12, 8, 8, 10, 11, 11, 11, 8, 11])


# --------------------------------------------------------------- 5. Quais
def _onglet_quais(ws, res):
    _titre(ws, "Planning quais — pic de véhicules simultanés à quai par site")
    _entete(ws, 3, ["Jour", "Site", "Pic véhicules simultanés à quai"])
    row = 4
    for jour in res:
        quais = res[jour].get("quais", {})
        for site, pic in sorted(quais.items(), key=lambda x: -x[1]):
            ws.cell(row=row, column=1, value=jour)
            ws.cell(row=row, column=2, value=site)
            c = ws.cell(row=row, column=3, value=pic); c.alignment = _CENTRE
            if pic >= 3:
                c.fill = _FILL_WARN
            row += 1
    ws.cell(row=row + 1, column=1,
            value="Les quais ne sont pas plafonnés ; cette mesure aide à repérer les sites tendus.")
    _largeurs(ws, [12, 28, 30])


# --------------------------------------------------------------- 6. Flux transportés
def _onglet_flux_transportes(ws, res):
    _titre(ws, "Flux transportés — rattachement flux → mission")
    cols = ["Jour", "Flux", "Origine", "Destination", "Contenant", "Quantité",
            "Type", "Mission", "Véhicule", "Sens", "Fonction support"]
    _entete(ws, 3, cols)
    row = 4
    for jour in res:
        # mission -> véhicule
        m2v = {}
        for p in res[jour]["postes"]:
            for m in p.missions:
                m2v[m.id] = p.id_vehicule
        for m in res[jour]["missions"]:
            for (fid, orig, dest, cont, qte) in m.composantes:
                ws.cell(row=row, column=1, value=jour)
                ws.cell(row=row, column=2, value=str(fid))
                ws.cell(row=row, column=3, value=orig)
                ws.cell(row=row, column=4, value=dest)
                ws.cell(row=row, column=5, value=cont)
                ws.cell(row=row, column=6, value=qte).alignment = _CENTRE
                ws.cell(row=row, column=7, value=m.propre_sale)
                ws.cell(row=row, column=8, value=m.id)
                ws.cell(row=row, column=9, value=m2v.get(m.id, "—"))
                ws.cell(row=row, column=10, value=m.sens)
                ws.cell(row=row, column=11, value=m.fonction_support)
                row += 1
    _largeurs(ws, [10, 8, 18, 18, 22, 9, 9, 10, 16, 10, 16])


# --------------------------------------------------------------- 7. Flux NON servis
def _onglet_flux_non_servis(ws, res):
    _titre(ws, "Flux NON servis — la solution n'est valide que si cet onglet est vide")
    cols = ["Jour", "Flux", "Origine", "Destination", "Contenant",
            "Pourquoi ce flux n'est pas planifié", "CONTRAINTE BLOQUANTE (à corriger)"]
    _entete(ws, 3, cols)
    row = 4
    vide = True
    for jour in res:
        for ns in res[jour].get("non_servis", []):
            vide = False
            ws.cell(row=row, column=1, value=jour)
            ws.cell(row=row, column=2, value=str(ns.get("flux_id", "")))
            ws.cell(row=row, column=3, value=ns.get("origine", ""))
            ws.cell(row=row, column=4, value=ns.get("destination", ""))
            ws.cell(row=row, column=5, value=ns.get("contenant", ""))
            c6 = ws.cell(row=row, column=6, value=ns.get("raison", "")); c6.alignment = _WRAP
            c7 = ws.cell(row=row, column=7, value=ns.get("contrainte", "")); c7.alignment = _WRAP
            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = _FILL_KO
            c6.fill = _FILL_KO
            c7.fill = _FILL_ALERTE; c7.font = Font(bold=True, color="FFFFFF")
            row += 1
    if vide:
        c = ws.cell(row=4, column=1, value="✓ Aucun flux non servi : solution VALIDE.")
        c.fill = _FILL_OK; c.font = Font(bold=True)
    _largeurs(ws, [10, 8, 18, 18, 22, 52, 46])


# --------------------------------------------------------------- 8. Contrôles
def _onglet_controles(ws, res, params):
    _titre(ws, "Contrôles des contraintes — vérifications automatiques")
    _entete(ws, 3, ["Jour", "Contrôle", "Résultat", "Détail"])
    row = 4
    rh = params.get("rh", {})
    amplitude = float(rh.get("amplitude_totale", 450)) if rh else 450
    pause_req = float(rh.get("pause", 30)) if rh else 30

    def ligne(jour, libelle, ok, detail):
        nonlocal row
        ws.cell(row=row, column=1, value=jour)
        ws.cell(row=row, column=2, value=libelle)
        c = ws.cell(row=row, column=3, value="OK" if ok else "ANOMALIE")
        c.fill = _FILL_OK if ok else _FILL_KO; c.alignment = _CENTRE
        ws.cell(row=row, column=4, value=detail).alignment = _WRAP
        row += 1

    for jour in res:
        postes = res[jour]["postes"]
        audit = res[jour].get("audit", {})
        # amplitude
        invalides = [p.id for p in postes if abs(p.amplitude - amplitude) > 1e-6]
        ligne(jour, f"Amplitude des postes = {amplitude:.0f} min", not invalides,
              "Tous conformes" if not invalides else f"Non conformes : {', '.join(invalides)}")
        # pause
        sans_pause = [p.id for p in postes
                      if sum(e.duree for e in p.etapes if e.type == "PAUSE") + 1e-6 < pause_req]
        ligne(jour, f"Pause obligatoire au dépôt ({pause_req:.0f} min)", not sans_pause,
              "Toutes prises au dépôt" if not sans_pause else f"Manquantes : {', '.join(sans_pause)}")
        # relève : pas de chevauchement sur un même véhicule
        from collections import defaultdict
        byv = defaultdict(list)
        for p in postes:
            byv[p.id_vehicule].append((p.h_debut, p.h_fin))
        chevau = []
        for v, sl in byv.items():
            sl.sort()
            for i in range(len(sl) - 1):
                if sl[i][1] > sl[i + 1][0]:
                    chevau.append(v)
        ligne(jour, "Relève : postes disjoints par véhicule", not chevau,
              "Aucun chevauchement" if not chevau else f"Conflits : {', '.join(set(chevau))}")
        # deadlines respectées
        retards = []
        for p in postes:
            for e in p.etapes:
                if e.type == "MISSION" and e.mission and e.h_fin > e.mission.h_deadline + 1e-6:
                    retards.append(e.mission.id)
        ligne(jour, "Fenêtres horaires (deadlines) respectées", not retards,
              "Toutes respectées" if not retards else f"En retard : {', '.join(retards)}")
        # flux non servis
        ns = res[jour].get("non_servis", [])
        ligne(jour, "Flux non servis", not ns,
              "Aucun" if not ns else f"{len(ns)} flux — voir onglet 7")
        ligne(jour, "Capacité maximale des véhicules respectée",
              not any("occupation véhicule" in a for a in audit.get("anomalies", [])),
              f"Maximum paramétré : {audit.get('taux_occupation_max_vehicule', 0):.0%}")
        ligne(jour, "Audit technique final", audit.get("nb_anomalies", 0) == 0,
              "Aucune anomalie" if audit.get("nb_anomalies", 0) == 0
              else " | ".join(audit.get("anomalies", [])))
    _largeurs(ws, [10, 38, 14, 70])


# --------------------------------------------------------------- 9. Indicateurs
def _onglet_indicateurs(ws, res):
    _titre(ws, "Indicateurs par jour")
    cols = ["Indicateur"] + list(res.keys())
    _entete(ws, 3, cols)
    lignes = [
        ("Flux servis (missions)", "nb_missions"),
        ("Postes (vacations)", "nb_postes"),
        ("Véhicules (flotte)", "nb_vehicules_total"),
        ("Pic véhicules simultanés", "pic_vehicules_simultanes"),
        ("Flux non servis", "nb_flux_non_servis"),
        ("Temps chargé (min)", "temps_charge_min"),
        ("Temps à vide (min)", "temps_vide_min"),
        ("Km à plein", "km_plein"),
        ("Km à vide", "km_vide"),
        ("Km total", "km_total"),
        ("Taux de km à vide (%)", "taux_km_vide"),
        ("Taux chargé / roulage (%)", "taux_charge_global"),
        ("Occupation moyenne (%)", "occupation_moyenne"),
        ("Postes occupés à moins de 80 %", "nb_postes_sous_80"),
        ("Solution valide", "solution_valide"),
        ("Pic véhicules à quai", "pic_quais"),
        ("Temps de calcul (s)", "temps_calcul_s"),
    ]
    for i, (lib, cle) in enumerate(lignes):
        r = 4 + i
        ws.cell(row=r, column=1, value=lib).font = Font(bold=True)
        for j, jour in enumerate(res.keys(), 2):
            ws.cell(row=r, column=j, value=res[jour]["metriques"].get(cle, "")).alignment = _CENTRE
    _largeurs(ws, [28] + [12] * len(res))
